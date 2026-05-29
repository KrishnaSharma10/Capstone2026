




import re
import json
import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from pymongo import MongoClient, UpdateOne

# ── Constants ──────────────────────────────────────────────────────────────────────

DAYS          = ['MON', 'TUE', 'WED', 'THU', 'FRI']
SLOTS_PER_DAY = 14

SKIP_SHEETS = {
    'pg time table 1',
    'pg time table1',
    'dlit',
}

_CONTINUATION_RE = re.compile(r'^LAB[-\s]?\d*$', re.IGNORECASE)

# NEW REGEX: Must be at least 5 chars long AND contain at least one digit
# Blocks faculty initials like "DMG" while allowing "UCSXX1"
_COURSE_RE = re.compile(r'^(?=.{5,})[A-Z]{2,}[A-Z0-9]*\d+', re.IGNORECASE)

DAY_MAP = {
    'MON': 'Monday', 'TUE': 'Tuesday', 'WED': 'Wednesday',
    'THU': 'Thursday', 'FRI': 'Friday',
}

SLOT_TIMES = {
    1: '8:00 AM',  2: '8:50 AM',  3: '9:40 AM',  4: '10:30 AM',
    5: '11:20 AM', 6: '12:10 PM', 7: '1:00 PM',  8: '1:50 PM',
    9: '2:40 PM', 10: '3:30 PM', 11: '4:20 PM', 12: '5:10 PM',
    13: '6:00 PM', 14: '6:50 PM',
}

TYPE_COLOR = {
    'lecture':  '#FFD700',
    'lab':      '#90EE90',
    'tutorial': '#ADD8E6',
    'elective': '#FFC0CB',
}


def slot_to_display(slot_str):
    parts = slot_str.split('-')
    return DAY_MAP.get(parts[0], parts[0]), SLOT_TIMES.get(int(parts[1]), parts[1])


def build_elective_flat_set(elective_data):
    codes = set()
    for basket_codes in elective_data.get('main-elective', {}).values():
        codes.update(basket_codes)
    return codes


# def parse_token(token):
#     """
#     Parse a single course token like 'UCS658L' -> ('UCS658', 'lab').
#     """
#     token = token.strip()
#     if not token:
#         return None
        
#     # Updated to handle more flexible alphanumeric base codes
#     m = re.match(r'^([A-Z0-9]+?)([LPT])?$', token, re.IGNORECASE)
#     if not m:
#         return None
#     code   = m.group(1).upper()
#     suffix = (m.group(2) or '').upper()
    
#     if not _COURSE_RE.match(code):
#         return None
        
#     typ = {'L': 'lecture', 'P': 'lab', 'T': 'tutorial'}.get(suffix, 'lecture')
#     return code, typ
def parse_token(token):
    """
    Parses a single course token. 
    FIX: Now allows trailing text (like GAMING LAB) but strictly validates the code.
    """
    token = token.strip()
    if not token:
        return None
        
    # Match the base code and the optional suffix (L/P/T).
    # Removed the '$' anchor so it doesn't fail if there's text after the code.
    m = re.match(r'^([A-Z0-9]+?)([LPT])?(?:\s|$)', token, re.IGNORECASE)
    if not m:
        return None
        
    code   = m.group(1).upper()
    suffix = (m.group(2) or '').upper()
    
    # Apply the strict 5-character + digit filter (Blocks DMG, SBH, etc.)
    if not _COURSE_RE.match(code):
        return None
        
    typ = {'L': 'lecture', 'P': 'lab', 'T': 'tutorial'}.get(suffix, 'lecture')
    return code, typ


def extract_codes_from_cell(raw, elective_flat_set):
    """
    Parses an Excel cell value.
    FIX 1: Handles multiline cells (UCT801P\nGAMING LAB) by taking the first line.
    FIX 2: Correctly tags single-code electives as 'elective'.
    """
    if not raw:
        return []
    
    # Split by newline and take the first line to ignore descriptive notes
    # Then remove spaces for the code processing
    raw_str = str(raw).split('\n')[0]
    s = raw_str.strip().replace(' ', '')
    
    if not s:
        return []

    # -- Combined class: UCS534+ULC664L ---------------------------------------
    if '+' in s:
        first_token = s.split('+')[0]
        parsed = parse_token(first_token)
        return [parsed] if parsed else []

    # -- Elective slash group: UCS751L/UCS752L --------------------------------
    if '/' in s:
        slash_tokens  = s.split('/')
        parsed_tokens = [p for p in (parse_token(t) for t in slash_tokens) if p]
        if not parsed_tokens:
            return []
        
        # If any code in the group is in the elective list, treat all as electives
        is_elective_group = any(code in elective_flat_set for code, _ in parsed_tokens)
        if is_elective_group:
            return [(code, 'elective') for code, _ in parsed_tokens]
        else:
            return [parsed_tokens[0]]

    # -- Single token (e.g., UCS635L or UCT801P) ------------------------------
    parsed = parse_token(s)
    if parsed:
        code, typ = parsed
        # GLOBAL FIX: Even if there is no slash, check if this code is an elective
        if code in elective_flat_set:
            return [(code, 'elective')]
        return [parsed]
        
    return []

# def extract_codes_from_cell(raw, elective_flat_set):
#     """
#     Parse a raw Excel cell value and return a list of (code, type) pairs.
#     """
#     if not raw:
#         return []
#     s = str(raw).strip().replace(' ', '')
#     if not s:
#         return []

#     # -- Combined class: UCS534+ULC664L (BUG C FIX) ---------------------------
#     if '+' in s:
#         first_token = s.split('+')[0]
#         parsed = parse_token(first_token)
#         return [parsed] if parsed else []

#     # -- Elective slash group: UCS751L/UCS752L or UCSXX2P/UCS658L -------------
#     if '/' in s:
#         slash_tokens  = s.split('/')
#         parsed_tokens = [p for p in (parse_token(t) for t in slash_tokens) if p]
#         if not parsed_tokens:
#             return []
        
#         is_elective_group = any(code in elective_flat_set for code, _ in parsed_tokens)
#         if is_elective_group:
#             return [(code, 'elective') for code, _ in parsed_tokens]
#         else:
#             return [parsed_tokens[0]]

#     # -- Single token ----------------------------------------------------------
#     parsed = parse_token(s)
#     if parsed:
#         code, typ = parsed
#         # GLOBAL FIX: Check if the single code is actually an elective
#         if code in elective_flat_set:
#             return [(code, 'elective')]
#         return [parsed]
#     return []


def is_continuation_of_previous(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    if _CONTINUATION_RE.match(s):
        return True
    if _COURSE_RE.match(s):
        return False
    return True


def slot_label(day_idx, slot_idx):
    return f"{DAYS[day_idx]}-{slot_idx + 1}"


# -- Merged Cell Lookup --------------------------------------------------------

def build_merge_lookup(sheet):
    lk = {}
    for mr in sheet.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                lk[(r, c)] = mr
    return lk


def read_cell(sheet, row, col, merge_lookup):
    mr = merge_lookup.get((row, col))
    if mr:
        return sheet.cell(mr.min_row, mr.min_col).value
    return sheet.cell(row, col).value


# -- Layout Detection ----------------------------------------------------------

def find_hours_cell(sheet):
    best, best_sum = None, float('inf')
    for row in sheet.iter_rows(max_row=15):
        for cell in row:
            if cell.value is not None and str(cell.value).strip().upper() in ('HOURS', 'HOUR'):
                s = cell.row + cell.column
                if s < best_sum:
                    best_sum = s
                    best     = (cell.row, cell.column)
    return best if best else (None, None)


def get_subgroup_columns(sheet, header_row, hours_col):
    SKIP_VALUES = {
        'HOURS', 'HOUR', 'DAY', 'DAYS', 'SR NO', 'SR.NO', 'BRANCH',
        'PRACTICAL', 'TUTORIAL', 'LECTURE', 'TOTAL', 'SIGNATURE', '',
        'ELECTRONICS', 'ELECTRICAL', 'COMPUTER SCIENCE', 'COMPUTER E',
        'COMPUTER ENGG', 'CIVIL ENGG', 'MECHANICAL ENGG', 'CHEMICAL ENGG',
        'BIOMEDICAL ENGG', 'BIOTECHNOLOGY ENGG',
    }
    sg_map    = {}
    seen_cols = set()
    for r in range(header_row, max(0, header_row - 4), -1):
        for c in range(hours_col + 1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val is None:
                continue
            s = str(val).strip()
            if not s or s.upper() in SKIP_VALUES:
                continue
            if s[0].isdigit() and 3 <= len(s) <= 6 and re.match(r'^[A-Z0-9]+$', s, re.IGNORECASE):
                if c not in seen_cols:
                    sg_map[s] = c
                    seen_cols.add(c)
    return sg_map


def get_sr_col(sheet, header_row):
    for c in range(1, 4):
        v = str(sheet.cell(header_row, c).value or '').strip().upper()
        if 'SR' in v or v in ('1', '2'):
            return c
    return 2


def get_first_data_row(sheet, header_row, sr_col):
    for r in range(header_row + 1, header_row + 8):
        v = sheet.cell(r, sr_col).value
        if v == 1 or v == '1':
            return r
    return header_row + 1


def detect_row_step(sheet, first_data_row, sr_col):
    v = sheet.cell(first_data_row + 1, sr_col).value
    return 1 if (v == 2 or v == '2') else 2


# -- Slot Iterator -------------------------------------------------------------

def iter_slot_rows(sheet, first_data_row, row_step, sr_col):
    row     = first_data_row
    max_row = sheet.max_row
    for day_idx in range(len(DAYS)):
        for slot_idx in range(SLOTS_PER_DAY):
            if row > max_row:
                return
            venue_row = row + 1 if row_step == 2 else row
            yield day_idx, slot_idx, row, venue_row
            row += row_step


# -- Per-Subgroup Builder ------------------------------------------------------

def build_subgroup_slots(sheet, sg_col, first_data_row, row_step,
                          sr_col, merge_lookup, elective_flat_set):
    slots_iter = list(iter_slot_rows(sheet, first_data_row, row_step, sr_col))
    entries    = []
    last_codes = []

    for (day_idx, slot_idx, code_row, venue_row) in slots_iter:
        raw_val = read_cell(sheet, code_row, sg_col, merge_lookup)

        if is_continuation_of_previous(raw_val):
            for (code, typ) in last_codes:
                entries.append({
                    'slot': slot_label(day_idx, slot_idx),
                    'code': code,
                    'type': typ,
                })
            continue

        last_codes = []
        if raw_val is None:
            continue
        s = str(raw_val).strip()
        if not s:
            continue

        parsed = extract_codes_from_cell(s, elective_flat_set)
        if not parsed:
            continue

        for (code, typ) in parsed:
            entries.append({
                'slot': slot_label(day_idx, slot_idx),
                'code': code,
                'type': typ,
            })
        last_codes = parsed

    return entries


# -- Aggregator ----------------------------------------------------------------

def aggregate(entries):
    result = {'lecture': {}, 'lab': {}, 'tutorial': {}, 'elective': {}}
    for e in entries:
        bucket = result[e['type']]
        if e['code'] not in bucket:
            bucket[e['code']] = []
        if e['slot'] not in bucket[e['code']]:
            bucket[e['code']].append(e['slot'])
    return result


# -- Elective subgroup list builder --------------------------------------------

def build_elective_subgrouplist(subgroups):
    result = {}
    for sg_name, sg_data in subgroups.items():
        elective_bucket = sg_data.get('elective', {})
        if not elective_bucket:
            continue
        slot_to_options = {}
        for code, slots in elective_bucket.items():
            for slot in slots:
                if slot not in slot_to_options:
                    slot_to_options[slot] = []
                if code not in slot_to_options[slot]:
                    slot_to_options[slot].append(code)
        if slot_to_options:
            result[sg_name] = [
                {'slot': slot, 'options': options}
                for slot, options in sorted(slot_to_options.items())
            ]
    return result


# -- Sheet & Workbook Processors -----------------------------------------------

def process_sheet(sheet, elective_flat_set):
    hours_row, hours_col = find_hours_cell(sheet)
    if hours_row is None:
        return {}, {}
    header_row = hours_row
    sg_map     = get_subgroup_columns(sheet, header_row, hours_col)
    if not sg_map:
        return {}, {}
    sr_col         = get_sr_col(sheet, header_row)
    first_data_row = get_first_data_row(sheet, header_row, sr_col)
    row_step       = detect_row_step(sheet, first_data_row, sr_col)
    merge_lookup   = build_merge_lookup(sheet)

    subgroup_dict       = {}
    course_to_subgroups = {}

    for sg_name, sg_col in sg_map.items():
        entries = build_subgroup_slots(
            sheet, sg_col, first_data_row, row_step,
            sr_col, merge_lookup, elective_flat_set)
        agg = aggregate(entries)
        subgroup_dict[sg_name] = agg
        for typ in ('lecture', 'lab', 'tutorial'):
            for code in agg[typ]:
                lst = course_to_subgroups.setdefault(code, [])
                if sg_name not in lst:
                    lst.append(sg_name)

    return subgroup_dict, course_to_subgroups


def should_process(sheet):
    found_hours = False
    for row in sheet.iter_rows(max_row=15, values_only=True):
        for cell in row:
            if cell and str(cell).strip().upper() in ('HOURS', 'HOUR'):
                found_hours = True
                break
        if found_hours:
            break
    if not found_hours:
        return False
    for row in sheet.iter_rows(max_row=15, values_only=True):
        for cell in row:
            if cell:
                s = str(cell).strip()
                if len(s) >= 3 and s[0].isdigit() and any(c.isalpha() for c in s):
                    return True
    return False


def process_workbook(path, elective_flat_set):
    wb        = load_workbook(path)
    subgroups = {}
    courses   = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if not should_process(sheet):
            continue
        sg_dict, course_map = process_sheet(sheet, elective_flat_set)
        for sg, data in sg_dict.items():
            if sg not in subgroups:
                subgroups[sg] = data
            else:
                for typ in ('lecture', 'lab', 'tutorial', 'elective'):
                    for code, slots in data[typ].items():
                        existing = subgroups[sg][typ].setdefault(code, [])
                        for s in slots:
                            if s not in existing:
                                existing.append(s)
        for code, sgs in course_map.items():
            existing = courses.setdefault(code, [])
            for sg in sgs:
                if sg not in existing:
                    existing.append(sg)
    return subgroups, courses


# -- MongoDB -------------------------------------------------------------------

def subgroups_to_mongo_format(subgroups):
    documents = []
    for sg_name, sg_data in subgroups.items():
        events = []
        for typ in ('lecture', 'lab', 'tutorial', 'elective'):
            color = TYPE_COLOR[typ]
            for code, slots in sg_data.get(typ, {}).items():
                for slot_str in slots:
                    try:
                        day, hour = slot_to_display(slot_str)
                    except Exception:
                        continue
                    events.append({
                        'day':         day,
                        'hour':        hour,
                        'subjectCode': code,
                        'subjectName': code,
                        'venue':       '',
                        'color':       color,
                    })
        documents.append({'subgroup': sg_name, 'data': events})
    return documents


def upsert_to_mongodb(documents, mongo_uri):
    try:
        client     = MongoClient(mongo_uri, serverSelectionTimeoutMS=10000)
        db         = client['capstone']
        collection = db['timeTableData2']

        deleted = collection.delete_many({})
        print(f"MongoDB: cleared {deleted.deleted_count} stale document(s)")

        operations = [
            UpdateOne(
                {'subgroup': doc['subgroup']},
                {'$set': doc},
                upsert=True,
            )
            for doc in documents
        ]
        if operations:
            result = collection.bulk_write(operations)
            print(f"MongoDB: upserted {result.upserted_count} new, "
                  f"modified {result.modified_count} existing subgroups")
        client.close()
    except Exception as e:
        print(f"MongoDB upsert failed: {e}")
        print("JSON files were still saved successfully.")


# -- PreprocessClass -----------------------------------------------------------

class PreprocessClass:

    def preprocessScriptFunc(self, path):
        print(f"Parsing: {path}")
        base = Path(__file__).parent

        elective_path = base / 'elective.json'
        if elective_path.exists():
            with open(elective_path, encoding='utf-8') as f:
                electives_data = json.load(f)
            elective_flat_set = build_elective_flat_set(electives_data)
            print(f"Loaded {len(elective_flat_set)} elective codes")
        else:
            print("Warning: elective.json not found -- all codes treated as regular")
            elective_flat_set = set()

        subgroups, courses = process_workbook(path, elective_flat_set)
        print(f"Parsed: {len(subgroups)} subgroups | {len(courses)} unique course codes")

        out_sg = base / 'subgroups.json'
        with open(out_sg, 'w', encoding='utf-8') as f:
            json.dump(subgroups, f, indent=2)
        print(f"Saved -> {out_sg}")

        out_c = base / 'courses.json'
        with open(out_c, 'w', encoding='utf-8') as f:
            json.dump(courses, f, indent=2)
        print(f"Saved -> {out_c}")

        elective_sglist = build_elective_subgrouplist(subgroups)
        out_esg = base / 'electives_subgrouplist.json'
        with open(out_esg, 'w', encoding='utf-8') as f:
            json.dump(elective_sglist, f, indent=2)
        print(f"Saved -> {out_esg} ({len(elective_sglist)} subgroups with elective slots)")

        mongo_uri = os.environ.get('MONGO_URI', '')
        if not mongo_uri:
            print("Warning: MONGO_URI not set -- skipping MongoDB update")
        else:
            print("Uploading to MongoDB...")
            documents = subgroups_to_mongo_format(subgroups)
            upsert_to_mongodb(documents, mongo_uri)
            run_post_upsert_tasks(mongo_uri)   # ← new line



# ── Post-Upsert: Version Save + Re-validation ─────────────────────────────────

from datetime import datetime, timezone

def save_timetable_version(db):
    """Save a UTC timestamp version string to metadata collection."""
    version_str = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    db.metadata.update_one(
        {"_id": "timetable_version"},
        {"$set": {"version": version_str, "updated_at": datetime.now(timezone.utc)}},
        upsert=True
    )
    print(f"[version] Timetable version saved: {version_str}")
    return version_str


def get_improvement_code(opted_courses):
    """
    opted_courses is [][]string in Go → stored as [[''UCS704''], ...]  in MongoDB.
    Handles that shape plus legacy string/dict shapes defensively.
    """
    if not opted_courses:
        return None
    first = opted_courses[0]
    if isinstance(first, list):          # ← correct shape: [["UCS704"]]
        return first[0] if first else None
    if isinstance(first, str):           # ← flat string fallback
        return first
    if isinstance(first, dict):          # ← object fallback
        return (first.get("course_code")
                or first.get("subjectCode")
                or first.get("code"))
    return None


def revalidate_approved_applications(db, version_str):
    """
    Re-check all stage 5 applications against the freshly uploaded timetable.
    If the improvement subject's slots now clash with anything in the new master
    timetable → set stage = 6, record timetable_version_flagged and clash_slots.
    """
    approved = list(db.applicationDetails.find({"stage": 5}))
    print(f"[revalidate] Checking {len(approved)} approved application(s)...")

    flagged = 0
    for app in approved:
        subgroup         = app.get("subgroup")
        new_tt           = app.get("new_time_table", [])
        opted_courses    = app.get("opted_courses", [])
        improvement_code = get_improvement_code(opted_courses)

        # Skip if any required field is missing
        if not subgroup or not new_tt or not improvement_code:
            continue

        # ── Step 1: slots the improvement subject occupies in the approved timetable
        improvement_slots = {
            (e["day"], e["hour"])
            for e in new_tt
            if e.get("subjectCode") == improvement_code
        }
        if not improvement_slots:
            continue

        # ── Step 2: all occupied slots in the NEW master timetable for this subgroup
        #           (exclude the improvement subject itself to avoid self-clash)
        master_doc = db.timeTableData2.find_one({"subgroup": subgroup})
        if not master_doc:
            continue

        master_slots = {
            (e["day"], e["hour"])
            for e in master_doc.get("data", [])
            if e.get("subjectCode") != improvement_code
        }

        # ── Step 3: intersection = actual new clashes
        clashes = improvement_slots & master_slots
        if not clashes:
            continue

        clash_list = [f"{day} {hour}" for day, hour in sorted(clashes)]
        db.applicationDetails.update_one(
            {"_id": app["_id"]},
            {"$set": {
                "stage":                      6,
                "timetable_version_flagged":  version_str,
                "clash_slots":                clash_list,
            }}
        )
        flagged += 1
        print(f"[revalidate] Flagged app {app.get('application_id')} "
              f"(subgroup {subgroup}) — clashes: {clash_list}")

    print(f"[revalidate] Done. {flagged}/{len(approved)} application(s) flagged.")


def run_post_upsert_tasks(mongo_uri):
    """Version save + re-validation in one call after timetable upsert."""
    try:
        client = MongoClient(mongo_uri, serverSelectionTimeoutMS=10000)
        db     = client["capstone"]
        version_str = save_timetable_version(db)
        revalidate_approved_applications(db, version_str)
        client.close()
    except Exception as e:
        print(f"[post-upsert] Error: {e}")

# -- CLI -----------------------------------------------------------------------

def main():
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        default = Path(__file__).parent / 'UG__PG_TIME_TABLE_JAN_TO_MAY_2026.xlsx'
        if default.exists():
            path = str(default)
        else:
            print("Usage: python preprocessScript.py <path_to_timetable.xlsx>")
            sys.exit(1)
    processor = PreprocessClass()
    processor.preprocessScriptFunc(path)


if __name__ == '__main__':
    main()

    
